# Import materials from .xlsx file

import csv
import json

from openpyxl import load_workbook


def import_xl(file: str) -> dict[str, list[str, float, str]]:

    wb = load_workbook(filename=file, data_only=True)
    sheet = wb.worksheets[0]

    data: dict[str, list[str, float, str]] = {}
    
    for row in sheet.iter_rows(min_row=1, max_col=12, values_only=True):
        if row[5] and row[5] != "Артикул":
            data[row[5]] = [row[0], row[11], ""]

    return data


def import_csv(file: str) -> dict[str, list[str, float, str]]:
    data: dict[str, list[str, float, str]] = {}
    with open(file, 'r', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        
        headers = next(reader)
        
        for row in reader:
            key = row[0]
            values = row[1:]
            data[key] = values
            data[key][1] = float(data[key][1])
        
    return data


def import_colors_json(file: str) -> dict[str, dict]:
    colors: dict[str, dict] = {}
    try:
        with open(file, 'r', encoding='utf-8') as file:
            colors = json.load(file)
    except FileNotFoundError:
        print("Файл не найден.")
    except json.JSONDecodeError:
        print("Ошибка декодирования JSON.")
    
    return colors


def main() -> None:
    print(import_xl("test.xlsx"))


if __name__ == "__main__":
    main()